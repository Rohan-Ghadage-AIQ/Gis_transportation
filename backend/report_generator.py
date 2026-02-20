"""
Generate downloadable delivery report from VRP results as Excel file
"""
import io
from datetime import datetime, timedelta
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def parse_time_str(time_str: str) -> int:
    """Convert time string like '09:30' (24-hour) to minutes since midnight"""
    try:
        parts = time_str.strip().split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return 0


def minutes_to_time_str(minutes: int) -> str:
    """Convert minutes since midnight to 24-hour time string like '09:30'"""
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours:02d}:{mins:02d}"


def determine_delivery_status(arrival_time_str: str, window_end: int) -> str:
    """
    Determine if delivery is on-time, late, or early
    
    Args:
        arrival_time_str: Arrival time like "09:30" (24-hour) or "N/A"
        window_end: Window end time in minutes since midnight
        
    Returns:
        Status: "ON_TIME", "LATE", "IN_BUFFER", or "UNKNOWN"
    """
    if arrival_time_str == "N/A" or not arrival_time_str:
        return "UNKNOWN"
    
    try:
        arrival_mins = parse_time_str(arrival_time_str)
        
        # If window_end is 0, assume any time is acceptable
        if window_end == 0:
            return "ON_TIME"
        
        diff = window_end - arrival_mins
        
        # If early by 60 mins or more -> IN_BUFFER (Safe)
        if diff >= 60:
            return "IN_BUFFER"
        # If early by 0-59 mins -> ON_TIME (Tight)
        elif diff >= 0:
            return "ON_TIME"
        # If late -> LATE
        else:
            return "LATE"
    except:
        return "UNKNOWN"


def generate_delivery_report(conn) -> bytes:
    """
    Generate comprehensive delivery report as Excel file
    
    Returns:
        Excel file as bytes
    """
    cursor = conn.cursor()
    
    # Get all deliveries with vehicle assignments
    query = """
        SELECT 
            s.station_id,
            s.vehicle_id,
            ST_X(s.geom) as longitude,
            ST_Y(s.geom) as latitude,
            s.parcel_weight,
            s.service_time,
            s.window_start,
            s.window_end,
            s.arrival_time,
            s.delivery_status
        FROM vector.station_node_map s
        WHERE s.vehicle_id IS NOT NULL
        ORDER BY s.vehicle_id, s.arrival_time
    """
    
    cursor.execute(query)
    deliveries = cursor.fetchall()
    
    # Calculate vehicle totals from route_geometries table
    cursor.execute("""
        SELECT 
            vehicle_id,
            ROUND((ST_Length(geom::geography)/1000)::numeric, 2) AS total_km
        FROM vector.route_geometries
        ORDER BY vehicle_id
    """)
    vehicle_distances = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Get vehicle shift times (must match vrp_solver.py)
    # Format: (start_time_str, duration_minutes)
    vehicle_shifts = [
        ("09:00", 540),  # V1: 09:00 - 18:00 (540, 1080) -> Dur: 540 (9h)
        ("09:00", 540),  # V2: 09:00 - 18:00 (540, 1080)
        ("07:00", 480),  # V3: 07:00 - 15:00 (420, 900) -> Dur: 480 (8h)
        ("07:00", 660),  # V4: 07:00 - 18:00 (420, 1080) -> Dur: 660 (11h)
        ("09:00", 480),  # V5: 09:00 - 17:00 (540, 1020) -> Dur: 480 (8h)
        ("08:00", 600),  # V6: 08:00 - 18:00 (480, 1080) -> Dur: 600 (10h)
        ("08:00", 780),  # V7: 08:00 - 21:00 (480, 1260) -> Dur: 780 (13h)
        ("07:00", 780),  # V8: 07:00 - 20:00 (420, 1200) -> Dur: 780 (13h)
        ("07:00", 720),  # V9: 07:00 - 19:00 (420, 1140) -> Dur: 720 (12h)
        ("08:00", 720)   # V10: 08:00 - 20:00 (480, 1200) -> Dur: 720 (12h)
    ]
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Report"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header
    headers = [
        "Vehicle ID", "Shift Start", "Shift End", "Total Distance (km)", 
        "Total Weight (kg)", "Parcel ID", "Parcel Weight (kg)", 
        "Service Time (min)", "Window End", "Arrival Time", 
        "Delivery Status", "On-Time Status"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    column_widths = [12, 12, 12, 18, 16, 15, 18, 18, 15, 15, 16, 16]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Group by vehicle to calculate totals
    vehicle_totals = {}
    for delivery in deliveries:
        vehicle_id = delivery[1]
        parcel_weight = delivery[4]
        if vehicle_id not in vehicle_totals:
            vehicle_totals[vehicle_id] = 0
        vehicle_totals[vehicle_id] += parcel_weight
    
    # Write data rows
    row_num = 2
    for delivery in deliveries:
        station_id = delivery[0]
        vehicle_id = delivery[1]
        parcel_weight = delivery[4]
        service_time = delivery[5]
        window_start = delivery[6]
        window_end = delivery[7]
        arrival_time = delivery[8] if delivery[8] else "N/A"
        delivery_status = delivery[9] if delivery[9] else "UNKNOWN"
        
        # Get vehicle totals
        total_km = vehicle_distances.get(vehicle_id, 0)
        total_weight_kg = vehicle_totals.get(vehicle_id, 0)
        
        # Get vehicle shift times
        shift_start, shift_duration = vehicle_shifts[vehicle_id - 1]
        shift_end_mins = parse_time_str(shift_start) + shift_duration
        shift_end = minutes_to_time_str(shift_end_mins)
        
        # Determine on-time status
        on_time_status = determine_delivery_status(arrival_time, window_end)
        
        # Format window_end
        window_end_str = minutes_to_time_str(window_end) if window_end > 0 else "Anytime"
        
        # Write row data
        row_data = [
            f"Vehicle {vehicle_id}",
            shift_start,
            shift_end,
            float(total_km),
            total_weight_kg,
            station_id,
            parcel_weight,
            service_time,
            window_end_str,
            arrival_time,
            delivery_status,
            on_time_status
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Color code on-time status
            if col_num == 12:  # On-Time Status column
                if value == "ON_TIME":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif value == "LATE":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
                elif value == "IN_BUFFER":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500")
        
        row_num += 1
    
        row_num += 1
    
    # ========================================
    # APPEND SHIFT-WISE SUMMARY (At Bottom)
    # ========================================
    
    # Define Shift Intervals (must match main.py SHIFTS)
    # 07:00-10:00 (420-600), 10:00-18:00 (600-1080), 18:00-21:00 (1080-1260)
    shift_intervals = {
        420: "07:00 - 10:00",
        600: "10:00 - 18:00",
        1080: "18:00 - 21:00"
    }

    # Group parcels by Shift (window_start)
    # deliveries index: 0:id, 4:weight, 5:service, 6:window_start, 7:window_end
    parcels_by_shift = {}
    for d in deliveries:
        w_start = d[6]
        if w_start not in parcels_by_shift:
            parcels_by_shift[w_start] = []
        parcels_by_shift[w_start].append(d)

    # Sort shifts chronologically
    sorted_starts = sorted(parcels_by_shift.keys())

    # Add spacing
    row_num += 3
    
    # Header for Summary Section
    header_cell = ws.cell(row=row_num, column=1, value="SHIFT-WISE PARCEL SUMMARY")
    header_cell.font = Font(bold=True, size=14, color="000000")
    row_num += 2

    # Iterate through each shift block
    for start_min in sorted_starts:
        shift_name = shift_intervals.get(start_min, f"Start: {minutes_to_time_str(start_min)}")
        
        # Shift Header
        shift_header_cell = ws.cell(row=row_num, column=1, value=f"Shift: {shift_name}")
        shift_header_cell.font = Font(bold=True, size=12, color="FFFFFF")
        shift_header_cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid") # Grey
        row_num += 1
        
        # Table Headers
        summary_headers = ["Parcel ID", "Parcel Weight (kg)", "Service Time (min)", "Window End"]
        for c, h in enumerate(summary_headers, 1):
             cell = ws.cell(row=row_num, column=c, value=h)
             cell.font = Font(bold=True)
             cell.border = border
             cell.alignment = Alignment(horizontal="center")
             # Light grey background for sub-headers
             cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") 
        row_num += 1
        
        # Data Rows
        for d in parcels_by_shift[start_min]:
             # d indices: 0:id, 4:weight, 5:service, 7:window_end
             p_id = d[0]
             p_weight = d[4]
             p_service = d[5]
             p_window_end = d[7]
             p_window_end_str = minutes_to_time_str(p_window_end) if p_window_end > 0 else "Anytime"
             
             ws.cell(row=row_num, column=1, value=p_id).border = border
             ws.cell(row=row_num, column=2, value=p_weight).border = border
             ws.cell(row=row_num, column=3, value=p_service).border = border
             ws.cell(row=row_num, column=4, value=p_window_end_str).border = border
             
             # Center align all
             for c in range(1, 5):
                 ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")
             
             row_num += 1
        
        row_num += 2 # Space between shift blocks

    # ========================================
    # CREATE UNASSIGNED PARCELS SHEET
    # ========================================
    
    # Query unassigned parcels (handle case where table doesn't exist yet)
    try:
        cursor.execute("""
            SELECT 
                station_id,
                reason,
                latitude,
                longitude,
                parcel_weight,
                window_end
            FROM vector.unassigned_parcels
            ORDER BY station_id
        """)
        unassigned = cursor.fetchall()
        print(f"DEBUG: Found {len(unassigned)} unassigned parcels for report")
    except Exception as e:
        # Table might not exist if no computation has been run yet
        print(f"Note: Could not query unassigned_parcels table: {e}")
        unassigned = []
    
    # Only create sheet if there are unassigned parcels
    if unassigned:
        ws_unassigned = wb.create_sheet("Unassigned Parcels")
        
        # Define header style for unassigned sheet (red theme)
        unassigned_header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        # Write headers
        unassigned_headers = [
            "Parcel ID", "Reason", "Latitude", "Longitude", 
            "Weight (kg)", "Window End"
        ]
        
        for col_num, header in enumerate(unassigned_headers, 1):
            cell = ws_unassigned.cell(row=1, column=col_num, value=header)
            cell.fill = unassigned_header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws_unassigned.column_dimensions['A'].width = 15  # Parcel ID
        ws_unassigned.column_dimensions['B'].width = 60  # Reason (wider for text)
        ws_unassigned.column_dimensions['C'].width = 15  # Latitude
        ws_unassigned.column_dimensions['D'].width = 15  # Longitude
        ws_unassigned.column_dimensions['E'].width = 15  # Weight
        ws_unassigned.column_dimensions['F'].width = 15  # Window End
        
        # Write data rows
        row_num = 2
        for parcel in unassigned:
            station_id, reason, lat, lon, weight, window_end = parcel
            window_end_str = minutes_to_time_str(window_end) if window_end > 0 else "Anytime"
            
            row_data = [station_id, reason, lat, lon, weight, window_end_str]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws_unassigned.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                
                # Left-align reason column for better readability
                if col_num == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row_num += 1
    
    cursor.close()
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
